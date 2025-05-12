import csv
import openpyxl
import argparse
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl import load_workbook

def main():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description='Convert CSV to XLSX and color code the whole mess. Do not include the extensions.')

    parser.add_argument(
        '-i', required=True,
        help='Postprocessed CSV file.')

    args = parser.parse_args()

    CSVname = args.i + ".csv"
    XLSXname = args.i + ".xlsx"


    wb = openpyxl.Workbook()
    ws = wb.active

    with open(CSVname) as f:
       reader = csv.reader(f, delimiter=',')
       for row in reader:
           ws.append(row)

    wb.save(XLSXname)


    wb = load_workbook(XLSXname)
    ws = wb.active


    for row in ws.iter_rows():
        for cell in row:
            try:
                if (cell.value) == "a":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

                if (cell.value) == "h":
                    cell.fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type = "solid")

                if (cell.value) == "b":
                    cell.fill = PatternFill(start_color="00cc00", end_color="00cc00", fill_type = "solid")

            except:
                pass
    wb.save(XLSXname)

if __name__ == "__main__":
    main()
